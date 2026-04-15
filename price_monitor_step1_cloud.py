"""
온라인 가격 모니터링 자동화 - GitHub Actions/클라우드용
네이버 쇼핑 API로 모델별 가격 수집 → 엑셀 파일 생성 → 이메일 발송(선택)

필수 환경변수:
  NAVER_CLIENT_ID
  NAVER_CLIENT_SECRET

선택 환경변수:
  ENABLE_VBA=false
  OPEN_FILE_AFTER_RUN=false
  EMAIL_FROM=
  EMAIL_TO=
  EMAIL_APP_PASSWORD=
"""

import os
import sys
import re
import smtplib
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from urllib.parse import quote, urlparse
from datetime import datetime
from email.message import EmailMessage
from collections import Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 환경변수 설정
# ──────────────────────────────────────────────

NAVER_CLIENT_ID = os.getenv("NAVER_CLIENT_ID")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET")
ENABLE_VBA = os.getenv("ENABLE_VBA", "false").lower() == "true"
OPEN_FILE_AFTER_RUN = os.getenv("OPEN_FILE_AFTER_RUN", "false").lower() == "true"

if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
    raise ValueError("NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 환경변수가 필요합니다.")

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PUMP_LIST_PATH = os.path.join(_BASE_DIR, "펌프리스트_시트.xlsx")
TEMPLATE_PATH = os.path.join(_BASE_DIR, "통합결과양식.xlsx")
OUTPUT_DIR = _BASE_DIR

DEALER_LIST_SUMMARY = [
    "서우기업","LG윌로펌프","경동기전","고강C&P","광진종합상사","굿펌프","나인티에스","대림상사",
    "대영상사","대풍상사","미라클YT펌프","삼흥E&P","서울종합펌프","서울펌프랜드","세광사",
    "수중모터주식회사","시대상사","에스에이치테크","엘지산업","윌로종합상사 영천","이조",
    "이피컴퍼니","전진","주식회사 리텍솔루션","주식회사 세종종합상사","카토건설중기","투빈",
    "퍼맥스","펌스","하경상사","국제티에스","광명상사","희성산업","펌프랜드","대산종합상사",
]
DEALER_LIST_DETAIL = [
    "서우기업","윌로펌프백화점","오아시스 펌프","서울피엠","펌프365","윌로펌프총판","펌프굿",
    "나인티에스","펌프파트너","이엔지마켓","따뜻함","펌프산업","워터테크","펌프닷컴",
    "윌로프로","샌프란시스코2","pump-damoa","서울펌프몰","윌로공식 SKS윌로펌프",
    "수중모터주식회사","시대몰","펌프프렌드","윌로펌프마켓","윌로종합","윌로펌프모터",
    "이피컴퍼니","EP COMPANY","펌프몰","윌로펌프온라인쇼핑몰","주식회사 리텍솔루션",
    "주식회사 세종종합상사","여담고","주식회사 투빈","펌프의 모든 것","펌프뱅크",
    "펌스pums","펌프탑","윈디샵","신세계몰","광명상사","신한일전기공식인증몰",
    "펌프마스터","대산공구",
]

EXCLUDE_KEYWORDS = [
    "GS", "GW", "GD", "GB", "GH", "GU",
    "PH-460A", "PB-88", "PH-260W", "PH-260A",
    "Faith Gathering", "PH-160W", "PH-460W"
]

# ──────────────────────────────────────────────
# 스타일 헬퍼
# ──────────────────────────────────────────────

THIN = Side(style="thin")
THICK = Side(style="medium")

def border_all(ws, row, col):
    ws.cell(row, col).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def header_style(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="C0C0C0")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def number_fmt(cell, fmt):
    cell.number_format = fmt

# ──────────────────────────────────────────────
# 펌프리스트 읽기
# ──────────────────────────────────────────────

def load_pump_list(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"필수 파일이 없습니다: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb["펌프리스트"]

    models = []
    outlets = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        alt, no, model, cost, msrp = (row[i] if i < len(row) else None for i in range(5))
        mall_name = row[8] if len(row) > 8 else None
        dealer = row[9] if len(row) > 9 else None

        if model and str(model).strip():
            models.append({
                "model": str(model).strip(),
                "alt_model": str(alt).strip() if alt else None,
                "cost": float(cost) if cost else 0.0,
                "msrp": float(msrp) if msrp else 0.0,
            })

        if mall_name and str(mall_name).strip():
            outlets.append((str(mall_name).strip(), str(dealer).strip() if dealer else ""))

    all_models = [m["model"] for m in models]
    return models, outlets, all_models

# ──────────────────────────────────────────────
# 네이버 쇼핑 API 호출
# ──────────────────────────────────────────────

def naver_search(keyword, client_id, client_secret):
    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret,
    }
    results = []
    encoded = quote(keyword)

    for page in range(1, 11):
        start = (page - 1) * 100 + 1
        url = (
            f"https://openapi.naver.com/v1/search/shop.json"
            f"?query={encoded}&display=100&start={start}"
        )
        resp = requests.get(url, headers=headers, timeout=10, verify=False)
        if resp.status_code != 200:
            print(f"  [API 오류] {resp.status_code} — {keyword}")
            break

        data = resp.json()
        total = int(data.get("total", 0))
        items = data.get("items", [])
        results.extend(items)

        if start + len(items) > total:
            break

    return results

# ──────────────────────────────────────────────
# 결과 필터링
# ──────────────────────────────────────────────

TAG_RE = re.compile(r"<[^>]+>")

def clean_title(title):
    return TAG_RE.sub("", title)

def should_include(title, target_model, all_models, exclude_kw):
    t = title.upper()

    for kw in exclude_kw:
        if kw.upper() in t:
            return False

    for m in all_models:
        if m != target_model and m.upper() in t:
            return False

    return target_model.upper() in t

# ──────────────────────────────────────────────
# 개별 모델 시트 생성
# ──────────────────────────────────────────────

def build_model_sheet(wb, model_info, items, all_models):
    sheet_name = model_info["model"].replace("/", ".")[:31]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    headers = ["판매처", "제목", "최저가", "인터넷 권장가", "물품대", "DC율"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        header_style(cell)

    cost = model_info["cost"]
    msrp = model_info["msrp"]
    out_row = 2

    for item in items:
        title = clean_title(item.get("title", ""))
        if not should_include(title, model_info["model"], all_models, EXCLUDE_KEYWORDS):
            continue

        lprice = int(item.get("real_price") or item.get("lprice", 0))
        mall = item.get("real_seller") or item.get("mallName", "")
        link = item.get("link", "")
        dc = (1 - ((lprice / 1.1) / cost)) * 100 if cost > 0 else None

        if msrp > 0 and lprice >= msrp:
            continue
        if dc is not None and dc >= 30:
            continue

        ws.cell(out_row, 1, mall)
        ws.cell(out_row, 2, title)
        ws.cell(out_row, 2).hyperlink = link
        ws.cell(out_row, 2).font = Font(color="0563C1", underline="single")

        if mall == "쿠팡":
            for c in range(1, 7):
                ws.cell(out_row, c).fill = PatternFill("solid", fgColor="FFFF99")

        ws.cell(out_row, 3, lprice)
        number_fmt(ws.cell(out_row, 3), "#,##0")

        ws.cell(out_row, 4, msrp)
        number_fmt(ws.cell(out_row, 4), "#,##0")

        ws.cell(out_row, 5, cost)
        number_fmt(ws.cell(out_row, 5), "#,##0")

        if dc is not None:
            ws.cell(out_row, 6, round(dc, 1))
            number_fmt(ws.cell(out_row, 6), '0.0"%"')
            if lprice < cost:
                ws.cell(out_row, 6).font = Font(color="FF0000")
            elif lprice > cost:
                ws.cell(out_row, 6).font = Font(color="0000FF")
        else:
            ws.cell(out_row, 6, "N/A")

        for c in range(1, 7):
            border_all(ws, out_row, c)

        out_row += 1

    last_row = out_row - 1
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:F{last_row}"
        data_rows = []
        for r in ws.iter_rows(min_row=2, max_row=last_row, values_only=False):
            data_rows.append([(c.value, c.number_format, c.font, c.hyperlink) for c in r])

        data_rows.sort(key=lambda r: (r[2][0] if isinstance(r[2][0], (int, float)) else float('inf')))

        for ri, row_data in enumerate(data_rows, 2):
            for ci, (val, fmt, fnt, hl) in enumerate(row_data, 1):
                cell = ws.cell(ri, ci, val)
                cell.number_format = fmt
                if fnt:
                    cell.font = Font(color=fnt.color, underline=fnt.underline, bold=fnt.bold)
                if hl:
                    cell.hyperlink = hl

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 14

    return ws, out_row - 1

# ──────────────────────────────────────────────
# 통합결과 시트 생성
# ──────────────────────────────────────────────

def build_summary_sheet(wb, models, outlets):
    if "통합결과" in wb.sheetnames:
        del wb["통합결과"]
    ws = wb.create_sheet("통합결과", 0)

    ws.merge_cells("A1:B2")
    ws["A1"] = "모델명 / 판매처&대리점"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="C0C0C0")
    ws["A1"].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for col_i, (mall, dealer) in enumerate(outlets, 3):
        ws.cell(1, col_i, mall).font = Font(bold=True)
        ws.cell(1, col_i).alignment = Alignment(horizontal="center")
        ws.cell(1, col_i).fill = PatternFill("solid", fgColor="C0C0C0")
        ws.cell(2, col_i, dealer).alignment = Alignment(horizontal="center")
        ws.cell(2, col_i).fill = PatternFill("solid", fgColor="E0E0E0")
        for r in [1, 2]:
            ws.cell(r, col_i).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        if mall == "쿠팡":
            ws.cell(1, col_i).fill = PatternFill("solid", fgColor="FFFF99")
            ws.cell(2, col_i).fill = PatternFill("solid", fgColor="FFFF99")

    base_row = 3
    for model_info in models:
        model = model_info["model"]
        sheet_name = model.replace("/", ".")[:31]

        ws.merge_cells(start_row=base_row, start_column=1, end_row=base_row + 2, end_column=1)
        ws.cell(base_row, 1, model)
        ws.cell(base_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(base_row, 1).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        for offset, label in enumerate(["최저가", "인터넷 권장가", "DC율"]):
            cell = ws.cell(base_row + offset, 2, label)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        if sheet_name in wb.sheetnames:
            model_ws = wb[sheet_name]
            mall_row_map = {}
            for r in model_ws.iter_rows(min_row=2):
                mall_nm = r[0].value
                if mall_nm and mall_nm not in mall_row_map:
                    link = r[1].hyperlink.target if r[1].hyperlink else None
                    mall_row_map[mall_nm] = {
                        "lprice": r[2].value,
                        "msrp": r[3].value,
                        "dc": r[5].value,
                        "link": link,
                    }

            for col_i, (mall, _) in enumerate(outlets, 3):
                if mall in mall_row_map:
                    row_data = mall_row_map[mall]
                    lprice = row_data["lprice"]
                    msrp = row_data["msrp"]
                    dc = row_data["dc"]
                    link = row_data["link"]

                    existing = ws.cell(base_row, col_i).value
                    if lprice and (not existing or (isinstance(existing, (int, float)) and lprice < existing)):
                        c = ws.cell(base_row, col_i, lprice)
                        c.number_format = "#,##0"
                        c.alignment = Alignment(horizontal="right")
                        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                        if link:
                            c.hyperlink = link
                            c.font = Font(color="0563C1", underline="single")
                    else:
                        c = ws.cell(base_row, col_i)
                        if not c.value and lprice:
                            c.value = lprice
                            c.number_format = "#,##0"
                            c.alignment = Alignment(horizontal="right")
                        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

                    c2 = ws.cell(base_row + 1, col_i, msrp if msrp else "")
                    c2.number_format = "#,##0"
                    c2.alignment = Alignment(horizontal="right")
                    c2.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

                    c3 = ws.cell(base_row + 2, col_i)
                    if isinstance(dc, (int, float)):
                        c3.value = dc / 100
                        c3.number_format = "0.0%"
                        if dc >= 25:
                            c3.fill = PatternFill("solid", fgColor="FF0101")
                        elif dc >= 22:
                            c3.fill = PatternFill("solid", fgColor="FF9696")
                        elif dc >= 20:
                            c3.fill = PatternFill("solid", fgColor="FF9601")
                        elif dc >= 17:
                            c3.fill = PatternFill("solid", fgColor="FFFF01")
                    else:
                        c3.value = dc or ""
                    c3.alignment = Alignment(horizontal="center")
                    c3.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                else:
                    for offset in range(3):
                        ws.cell(base_row + offset, col_i).border = Border(
                            left=THIN, right=THIN, top=THIN, bottom=THIN
                        )

        for col_i in range(1, len(outlets) + 3):
            cell = ws.cell(base_row + 2, col_i)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=THICK
            )
        base_row += 3

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for col_i in range(3, len(outlets) + 3):
        ws.column_dimensions[get_column_letter(col_i)].width = 16

    for col_i in range(1, len(outlets) + 3):
        cell = ws.cell(2, col_i)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=THICK
        )

    ws.freeze_panes = "C3"
    return ws

# ──────────────────────────────────────────────
# 위반 정리 시트 생성
# ──────────────────────────────────────────────

def build_violation_sheet(wb):
    today = datetime.now().strftime("%Y.%m.%d")

    if "통합결과" not in wb.sheetnames:
        return

    ws_result = wb["통합결과"]
    last_col = min(ws_result.max_column, 108)
    last_data_row = ws_result.max_row

    ws_result.cell(1, 109, "최저가")
    ws_result.cell(2, 109, "최저가격")
    ws_result.cell(2, 110, "c2 or WILO")

    row_i = 3
    while row_i <= last_data_row:
        model_name = ws_result.cell(row_i, 1).value
        if not model_name:
            row_i += 3
            continue

        min_price = None
        min_seller = None
        min_msrp = None
        min_dc = None

        for col_i in range(3, last_col + 1):
            if col_i in (109, 110):
                continue
            lprice = ws_result.cell(row_i, col_i).value
            msrp = ws_result.cell(row_i + 1, col_i).value
            dc = ws_result.cell(row_i + 2, col_i).value
            mall = ws_result.cell(1, col_i).value

            if lprice and isinstance(lprice, (int, float)) and lprice > 0:
                if min_price is None or lprice < min_price:
                    min_price = lprice
                    min_seller = mall
                    min_msrp = msrp
                    min_dc = dc

        if min_price:
            ws_result.cell(row_i, 109, min_price)
            ws_result.cell(row_i + 1, 109, min_msrp or 0)
            ws_result.cell(row_i + 2, 109, (min_dc * 100) if min_dc and isinstance(min_dc, float) and min_dc < 1 else (min_dc or 0))
            ws_result.cell(row_i, 110, min_seller or "")
            ws_result.cell(row_i, 109).number_format = "#,##0"
            ws_result.cell(row_i + 1, 109).number_format = "#,##0"
            ws_result.cell(row_i + 2, 109).number_format = "0.00"
        else:
            ws_result.cell(row_i, 109, 0)
            ws_result.cell(row_i + 1, 109, 0)
            ws_result.cell(row_i + 2, 109, 0)
            ws_result.cell(row_i, 110, "wilo")

        for _r in range(row_i, row_i + 3):
            for _c in [109, 110]:
                ws_result.cell(_r, _c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        row_i += 3

    ws_result.cell(168, 2, "평균 dc 율")
    for col_i in range(3, last_col + 1):
        if col_i in (109, 110):
            continue
        dc_vals = []
        r = 5
        while r <= last_data_row:
            dc = ws_result.cell(r, col_i).value
            if dc and isinstance(dc, (int, float)) and dc != 0:
                dc_vals.append(dc if dc < 1 else dc / 100)
            r += 3
        avg = sum(dc_vals) / len(dc_vals) if dc_vals else 0
        cell = ws_result.cell(168, col_i, avg)
        cell.number_format = "0.0%"

    ws_result.cell(169, 2, "권장가 미만 개수")
    for col_i in range(3, last_col + 1):
        if col_i in (109, 110):
            continue
        count = 0
        r = 3
        while r <= last_data_row:
            lprice = ws_result.cell(r, col_i).value
            if lprice and isinstance(lprice, (int, float)) and lprice > 0:
                count += 1
            r += 3
        ws_result.cell(169, col_i, count)

    for r in [168, 169]:
        ws_result.cell(r, 2).font = Font(bold=True)
        ws_result.cell(r, 2).fill = PatternFill("solid", fgColor="C0C0C0")
        ws_result.cell(r, 2).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for col_i in range(3, last_col + 1):
        if col_i in (109, 110):
            continue
        for r in [168, 169]:
            ws_result.cell(r, col_i).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for r in [1, 2]:
        for c in [109, 110]:
            ws_result.cell(r, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            ws_result.cell(r, c).fill = PatternFill("solid", fgColor="C0C0C0")
            ws_result.cell(r, c).font = Font(bold=True)
            ws_result.cell(r, c).alignment = Alignment(horizontal="center")

    if "C2점 권장가 위반 정리" in wb.sheetnames:
        ws_target = wb["C2점 권장가 위반 정리"]
        target_row = ws_target.max_row + 3
    else:
        summary_idx = wb.sheetnames.index("통합결과") if "통합결과" in wb.sheetnames else 0
        ws_target = wb.create_sheet("C2점 권장가 위반 정리", summary_idx + 2)
        target_row = 1

    ws_target.cell(target_row, 1, "진행 날짜: %s" % today).font = Font(bold=True)
    target_row += 1

    for c, h in enumerate(["판매처", "대리점", "평균 DC율", "권장가 미만 개수"], 1):
        cell = ws_target.cell(target_row, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="C0C0C0")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center")
    target_row += 1
    data_end_row = target_row - 1

    for col_i in range(3, last_col + 1):
        if col_i in (109, 110):
            continue
        count_val = ws_result.cell(169, col_i).value
        if count_val and isinstance(count_val, (int, float)) and count_val > 0:
            mall_name = ws_result.cell(1, col_i).value
            dealer_name = ws_result.cell(2, col_i).value
            avg_dc = ws_result.cell(168, col_i).value or 0
            ws_target.cell(target_row, 1, mall_name or "")
            ws_target.cell(target_row, 2, dealer_name or "")
            ws_target.cell(target_row, 3, avg_dc).number_format = "0.0%"
            ws_target.cell(target_row, 4, int(count_val))
            for c in range(1, 5):
                ws_target.cell(target_row, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if dealer_name in DEALER_LIST_SUMMARY:
                for c in range(1, 5):
                    ws_target.cell(target_row, c).fill = PatternFill("solid", fgColor="FFFF00")
            target_row += 1
    data_end_row = target_row - 1

    extract_row = data_end_row + 3
    for c, h in enumerate(["모델명", "최저가", "DC율", "대리점명"], 1):
        cell = ws_target.cell(extract_row, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="C0C0C0")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center")
    extract_row += 1

    row_i = 3
    while row_i <= last_data_row:
        model_name = ws_result.cell(row_i, 1).value
        if not model_name:
            row_i += 3
            continue
        seller = ws_result.cell(row_i, 110).value or ""
        lprice = ws_result.cell(row_i, 109).value
        dc = ws_result.cell(row_i + 2, 109).value

        if (seller and seller.lower() not in ("wilo", "c2", "") and lprice and isinstance(lprice, (int, float)) and lprice > 0):
            ws_target.cell(extract_row, 1, model_name)
            ws_target.cell(extract_row, 2, lprice).number_format = "#,##0"
            if dc and isinstance(dc, (int, float)):
                dc_pct = dc / 100 if dc > 1 else dc
                ws_target.cell(extract_row, 3, dc_pct).number_format = "0.0%"
            ws_target.cell(extract_row, 4, seller)
            for c in range(1, 5):
                ws_target.cell(extract_row, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if seller in DEALER_LIST_DETAIL:
                for c in range(1, 5):
                    ws_target.cell(extract_row, c).fill = PatternFill("solid", fgColor="FFFF00")
            extract_row += 1
        row_i += 3

    ws_target.column_dimensions["A"].width = 20
    ws_target.column_dimensions["B"].width = 20
    ws_target.column_dimensions["C"].width = 12
    ws_target.column_dimensions["D"].width = 12
    print("  C2점 권장가 위반 정리 시트 생성 완료")

# ──────────────────────────────────────────────
# VBA 삽입 / 이메일 발송
# ──────────────────────────────────────────────

def insert_vba_macros(output_path):
    try:
        import win32com.client as win32
    except ImportError:
        print("  [참고] pywin32 미설치 — VBA 삽입 건너뜀")
        return output_path

    bas_path = os.path.join(_BASE_DIR, "매크로모음.bas")
    if not os.path.exists(bas_path):
        print("  [참고] 매크로모음.bas 파일 없음 — VBA 삽입 건너뜀")
        return output_path

    with open(bas_path, "r", encoding="utf-8") as bf:
        raw = bf.read()
    vba_code = raw.split("\n", 1)[1] if raw.startswith("Attribute") else raw
    xlsm_path = output_path.replace(".xlsx", ".xlsm")

    try:
        xl = win32.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(os.path.abspath(output_path))
        vba_module = wb.VBProject.VBComponents.Add(1)
        vba_module.Name = "매크로모음"
        vba_module.CodeModule.AddFromString(vba_code.strip())
        wb.SaveAs(os.path.abspath(xlsm_path), FileFormat=52)
        wb.Close(SaveChanges=False)
        xl.Quit()

        if os.path.exists(output_path) and output_path != xlsm_path:
            os.remove(output_path)

        print("  VBA 매크로 삽입 완료")
        return xlsm_path
    except Exception as e:
        print("  [참고] VBA 삽입 실패:", str(e))
        return output_path

def send_email_with_attachment(file_path):
    smtp_user = os.getenv("EMAIL_FROM")
    smtp_to = os.getenv("EMAIL_TO")
    smtp_pass = os.getenv("EMAIL_APP_PASSWORD")

    if not smtp_user or not smtp_to or not smtp_pass:
        print("[참고] 이메일 환경변수가 없어 메일 발송은 건너뜁니다.")
        return

    msg = EmailMessage()
    msg["Subject"] = f"온라인 가격 모니터링 결과 - {datetime.now().strftime('%Y-%m-%d')}"
    msg["From"] = smtp_user
    msg["To"] = smtp_to
    msg.set_content("자동 실행된 온라인 가격 모니터링 결과 파일을 첨부합니다.")

    filename = os.path.basename(file_path)
    subtype = "vnd.ms-excel.sheet.macroEnabled.12" if filename.lower().endswith(".xlsm") else "vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    with open(file_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype=subtype,
            filename=filename,
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(smtp_user, smtp_pass)
        smtp.send_message(msg)

    print("✅ 이메일 발송 완료")

# ──────────────────────────────────────────────
# 메인 실행
# ──────────────────────────────────────────────

def main():
    today = datetime.now().strftime("%Y.%m.%d")
    base_name = f"{today}_온라인_모니터링_파일"
    output_path = os.path.join(OUTPUT_DIR, f"{base_name}.xlsx")
    version = 2
    while os.path.exists(output_path):
        output_path = os.path.join(OUTPUT_DIR, f"{base_name}_v{version}.xlsx")
        version += 1

    print(f"[{today}] 온라인 가격 모니터링 시작\n")

    try:
        sys.path.insert(0, _BASE_DIR)
        from crawler_step2 import enrich_items_with_seller
        USE_CRAWLER = True
        print("  [2단계] 오픈마켓 크롤러 활성화\n")
    except ImportError:
        USE_CRAWLER = False
        print("  [참고] crawler_step2.py 없거나 관련 의존성 미설치 → 오픈마켓 판매자 자동추출 건너뜀\n")

    print("▶ 펌프리스트 로드 중...")
    models, outlets, all_models = load_pump_list(PUMP_LIST_PATH)
    print(f"  모델 수: {len(models)}개 | 판매처 수: {len(outlets)}개\n")

    wb = Workbook()
    wb.remove(wb.active)

    for i, model_info in enumerate(models, 1):
        model = model_info["model"]
        alt = model_info.get("alt_model")
        print(f"  [{i:02d}/{len(models)}] {model} 검색 중...", end=" ")

        items = naver_search(model, NAVER_CLIENT_ID, NAVER_CLIENT_SECRET)
        if alt:
            items += naver_search(alt, NAVER_CLIENT_ID, NAVER_CLIENT_SECRET)
        print(f"{len(items)}건 수집", end=" → ")

        if USE_CRAWLER:
            cost = model_info["cost"]
            msrp = model_info["msrp"]
            OPEN_MARKETS = {"쿠팡", "11번가", "G마켓", "옥션", "Gmarket", "Auction"}
            pre_filtered = []
            for item in items:
                title = clean_title(item.get("title", ""))
                if not should_include(title, model, all_models, EXCLUDE_KEYWORDS):
                    continue
                lp = int(item.get("lprice", 0))
                dc = (1 - ((lp / 1.1) / cost)) * 100 if cost > 0 else None
                if msrp > 0 and lp >= msrp:
                    continue
                if dc is not None and dc >= 30:
                    continue
                pre_filtered.append(item)

            open_targets = [it for it in pre_filtered if it.get("mallName", "") in OPEN_MARKETS][:100]
            if open_targets:
                open_targets = enrich_items_with_seller(open_targets)
                link_map = {it["link"]: it for it in open_targets}
                for item in items:
                    if item.get("link") in link_map:
                        item["real_seller"] = link_map[item["link"]].get("real_seller")
                        item["real_price"] = link_map[item["link"]].get("real_price")

        _, data_rows = build_model_sheet(wb, model_info, items, all_models)
        print(f"필터 후 {data_rows}행")

    print("\n▶ 통합결과 시트 생성 중...")
    build_summary_sheet(wb, models, outlets)

    OPEN_MARKET_LIST = {"쿠팡", "11번가", "G마켓", "옥션", "Gmarket", "Auction"}
    market_items = []
    seen_links = set()

    for sheet_name in wb.sheetnames:
        if sheet_name in ("통합결과", "오픈마켓확인"):
            continue
        ws_tmp = wb[sheet_name]
        for row in ws_tmp.iter_rows(min_row=2):
            if len(row) < 6:
                continue
            mall_val = row[0].value
            title_val = row[1].value
            lprice_val = row[2].value
            msrp_val = row[3].value
            dc_val = row[5].value
            hl = row[1].hyperlink
            if hl:
                link = hl.target
                if not link or link in seen_links:
                    continue
                if not (msrp_val and lprice_val and lprice_val < msrp_val):
                    continue

                host = urlparse(link).netloc.lower()
                if "coupang" in host:
                    detected_mall = "쿠팡"
                elif "11st" in host:
                    detected_mall = "11번가"
                elif "gmarket" in host or "link.gmarket" in host:
                    detected_mall = "G마켓"
                elif "auction" in host or "link.auction" in host:
                    detected_mall = "옥션"
                else:
                    continue

                market_items.append({
                    "mall": detected_mall,
                    "seller": mall_val,
                    "model": sheet_name,
                    "title": title_val,
                    "lprice": lprice_val,
                    "msrp": msrp_val,
                    "dc": dc_val,
                    "link": link,
                    "need_manual": detected_mall == "쿠팡",
                })
                seen_links.add(link)

    if "오픈마켓확인" in wb.sheetnames:
        del wb["오픈마켓확인"]
    summary_idx = wb.sheetnames.index("통합결과") if "통합결과" in wb.sheetnames else 0
    ws_mk = wb.create_sheet("오픈마켓확인", summary_idx + 1)

    headers_mk = ["쇼핑몰", "모델명", "상품명", "최저가", "권장가", "DC율", "판매자", "링크"]
    for c, h in enumerate(headers_mk, 1):
        cell = ws_mk.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="FFCC00")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    MALL_COLORS = {
        "쿠팡": "FFEEEE",
        "11번가": "EEF4FF",
        "G마켓": "EEFFF0",
        "옥션": "FFF8EE",
    }

    if market_items:
        order = {"쿠팡": 0, "11번가": 1, "G마켓": 2, "옥션": 3}
        market_items.sort(key=lambda x: (order.get(x["mall"], 9), x["model"]))

        for r_i, item in enumerate(market_items, 2):
            mall = item["mall"]
            color = MALL_COLORS.get(mall, "FFFFFF")

            ws_mk.cell(r_i, 1, mall)
            ws_mk.cell(r_i, 2, item["model"])
            ws_mk.cell(r_i, 3, item["title"])
            ws_mk.cell(r_i, 3).hyperlink = item["link"]
            ws_mk.cell(r_i, 3).font = Font(color="0563C1", underline="single")
            ws_mk.cell(r_i, 4, item["lprice"]).number_format = "#,##0"
            ws_mk.cell(r_i, 5, item["msrp"]).number_format = "#,##0"

            dc = item["dc"]
            if isinstance(dc, (int, float)):
                dc_val = dc / 100 if dc > 1 else dc
                ws_mk.cell(r_i, 6, dc_val).number_format = "0.0%"

            if item["need_manual"]:
                ws_mk.cell(r_i, 7, "← 직접 입력")
                ws_mk.cell(r_i, 7).fill = PatternFill("solid", fgColor="FFFF99")
            else:
                ws_mk.cell(r_i, 7, item["seller"])

            ws_mk.cell(r_i, 8, item["link"])
            ws_mk.cell(r_i, 8).hyperlink = item["link"]
            ws_mk.cell(r_i, 8).font = Font(color="0563C1", underline="single")

            for c in range(1, 9):
                ws_mk.cell(r_i, c).fill = PatternFill("solid", fgColor=color)
                ws_mk.cell(r_i, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if item["need_manual"]:
                ws_mk.cell(r_i, 7).fill = PatternFill("solid", fgColor="FFFF99")

        ws_mk.column_dimensions["A"].width = 10
        ws_mk.column_dimensions["B"].width = 16
        ws_mk.column_dimensions["C"].width = 50
        ws_mk.column_dimensions["D"].width = 12
        ws_mk.column_dimensions["E"].width = 12
        ws_mk.column_dimensions["F"].width = 10
        ws_mk.column_dimensions["G"].width = 22
        ws_mk.column_dimensions["H"].width = 55
        ws_mk.freeze_panes = "A2"

        mall_counts = Counter(item["mall"] for item in market_items)
        print(f"\n▶ 오픈마켓확인 시트 생성 완료: 총 {len(market_items)}개")
        for mall, cnt in sorted(mall_counts.items()):
            manual = " (← 판매자 수동입력 필요)" if mall == "쿠팡" else ""
            print(f"  {mall}: {cnt}개{manual}")
    else:
        ws_mk.cell(2, 1, "권장가 이하 오픈마켓 항목 없음")
        print("\n  오픈마켓 권장가 이하 항목 없음")

    wb.save(output_path)
    print(f"\n✅ 저장 완료: {output_path}")

    if ENABLE_VBA:
        print("\n▶ VBA 매크로 삽입 중...")
        final_path = insert_vba_macros(output_path)
        output_path = final_path or output_path

    send_email_with_attachment(output_path)

    if OPEN_FILE_AFTER_RUN:
        try:
            os.startfile(output_path)
        except Exception as e:
            print(f"[참고] 파일 자동 열기 실패: {e}")

    print(f"\n✅ 최종 완료! 저장 위치: {output_path}")
    return output_path

if __name__ == "__main__":
    main()
